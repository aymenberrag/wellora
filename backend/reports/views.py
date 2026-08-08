from datetime import datetime

from django.db.models import Avg, Count, Q, Sum, Max
from django.db.models.functions import TruncMonth
from django.http import FileResponse, HttpResponse

from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView

from core.permissions import ReportPermission, scope_queryset_for_user

from reports.utils.intervention_pdf import generate_intervention_pdf
from reports.utils.field_excel import create_field_workbook
from reports.utils.field_pdf import generate_field_pdf
from reports.utils.measurement_excel import create_measurement_workbook
from reports.utils.measurement_pdf import generate_measurement_pdf
from reports.utils.maintenance_excel import create_maintenance_workbook
from reports.utils.maintenance_pdf import generate_maintenance_pdf
from reports.utils.production_excel import create_production_workbook
from reports.utils.well_excel import create_well_workbook
from reports.utils.well_pdf import generate_well_pdf
from reports.utils.well_test_pdf import export_well_test_pdf
from .utils.production_pdf import generate_production_pdf
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from production.models import Production
from maintenance.models import Maintenance
from measurements.models import WellMeasurement
from fields.models import Field
from well_tests.models import WellTest
from wells.models import Well
from interventions.models import WellIntervention

class ProductionReportView(APIView):
    permission_classes = [ReportPermission]

    def build_report(self, request):
        queryset = scope_queryset_for_user(
            request.user,
            Production.objects.select_related(
                "well",
                "well__field",
                "well__operator",
            ),
            "production",
        )

        field_id = request.query_params.get("field")
        well_id = request.query_params.get("well")
        date_from = request.query_params.get("date_from")
        date_to = request.query_params.get("date_to")

        # -----------------------------
        # Filters
        # -----------------------------

        if field_id:
            queryset = queryset.filter(
                well__field_id=field_id
            )

        if well_id:
            queryset = queryset.filter(
                well_id=well_id
            )

        if date_from:
            try:
                date_from = datetime.strptime(
                    date_from,
                    "%Y-%m-%d"
                ).date()

                queryset = queryset.filter(
                    production_date__gte=date_from
                )

            except ValueError:
                raise ValueError(
                    "Invalid date_from format. Use YYYY-MM-DD."
                )

        if date_to:
            try:
                date_to = datetime.strptime(
                    date_to,
                    "%Y-%m-%d"
                ).date()

                queryset = queryset.filter(
                    production_date__lte=date_to
                )

            except ValueError:
                raise ValueError(
                    "Invalid date_to format. Use YYYY-MM-DD."
                )

        # -----------------------------
        # Summary
        # -----------------------------

        summary = queryset.aggregate(
            total_oil=Sum("oil_production"),
            total_gas=Sum("gas_production"),
            total_water=Sum("water_production"),
            average_oil=Avg("oil_production"),
            average_gas=Avg("gas_production"),
            average_water=Avg("water_production"),
            total_records=Count("id"),
        )

        # Convert None to 0
        for key in summary:
            if summary[key] is None:
                summary[key] = 0

        # -----------------------------
        # Daily production
        # -----------------------------

        daily_data = (
            queryset
            .values("production_date")
            .annotate(
                oil=Sum("oil_production"),
                gas=Sum("gas_production"),
                water=Sum("water_production"),
            )
            .order_by("production_date")
        )

        daily_production = []

        for item in daily_data:
            daily_production.append(
                {
                    "date": item["production_date"],
                    "oil": item["oil"] or 0,
                    "gas": item["gas"] or 0,
                    "water": item["water"] or 0,
                }
            )

        # -----------------------------
        # Top producing wells
        # -----------------------------

        top_wells_data = (
            queryset
            .values(
                "well_id",
                "well__code",
                "well__name",
            )
            .annotate(
                total_oil=Sum("oil_production"),
                total_gas=Sum("gas_production"),
                total_water=Sum("water_production"),
            )
            .order_by("-total_oil")[:10]
        )

        top_wells = []

        for item in top_wells_data:
            top_wells.append(
                {
                    "well_id": item["well_id"],
                    "well_code": item["well__code"],
                    "well_name": item["well__name"],
                    "oil": item["total_oil"] or 0,
                    "gas": item["total_gas"] or 0,
                    "water": item["total_water"] or 0,
                }
            )

        # -----------------------------
        # Final report data
        # -----------------------------

        return {
            "report": "Production Report",

            "filters": {
                "field": field_id,
                "well": well_id,
                "date_from": request.query_params.get(
                    "date_from"
                ),
                "date_to": request.query_params.get(
                    "date_to"
                ),
            },

            "summary": summary,

            "daily_production": daily_production,

            "top_wells": top_wells,
        }

    def get(self, request):
        try:
            report_data = self.build_report(request)

        except ValueError as error:
            return Response(
                {
                    "error": str(error)
                },
                status=400,
            )

        return Response(report_data)

    def pdf(self, request):
        try:
            report_data = self.build_report(request)

        except ValueError as error:
            return Response(
                {
                    "error": str(error)
                },
                status=400,
            )

        pdf_buffer = generate_production_pdf(
            report_data
        )

        return FileResponse(
            pdf_buffer,
            as_attachment=True,
            filename="production-report.pdf",
            content_type="application/pdf",
        )

class ProductionReportPDFView(ProductionReportView):
    def get(self, request):
        return super().pdf(request)


class ProductionReportExcelView(ProductionReportView):
    def get(self, request):
        try:
            report_data = self.build_report(request)
        except ValueError as error:
            return Response({"error": str(error)}, status=400)

        workbook = create_production_workbook(report_data)
        response = HttpResponse(
            content_type=(
                "application/vnd.openxmlformats-"
                "officedocument.spreadsheetml.sheet"
            )
        )
        response[
            "Content-Disposition"
        ] = 'attachment; filename="production_report.xlsx"'
        workbook.save(response)
        return response



class WellTestReportView(APIView):
    permission_classes = [ReportPermission]

    def get(self, request):
        queryset = scope_queryset_for_user(
            request.user,
            WellTest.objects.select_related(
                "well",
                "well__field",
            ),
            "well_tests",
        )

        # -----------------------------
        # Filters
        # -----------------------------

        field_id = request.query_params.get("field")
        well_id = request.query_params.get("well")
        date_from = request.query_params.get("date_from")
        date_to = request.query_params.get("date_to")

        if field_id:
            queryset = queryset.filter(
                well__field_id=field_id
            )

        if well_id:
            queryset = queryset.filter(
                well_id=well_id
            )

        if date_from:
            queryset = queryset.filter(
                test_date__gte=date_from
            )

        if date_to:
            queryset = queryset.filter(
                test_date__lte=date_to
            )

        # -----------------------------
        # Summary
        # -----------------------------

        summary = queryset.aggregate(
            total_tests=Count("id"),

            average_oil_rate=Avg(
                "oil_rate"
            ),

            average_gas_rate=Avg(
                "gas_rate"
            ),

            average_water_rate=Avg(
                "water_rate"
            ),

            average_wellhead_pressure=Avg(
                "wellhead_pressure"
            ),

            average_bottomhole_pressure=Avg(
                "bottomhole_pressure"
            ),

            average_water_cut=Avg(
                "water_cut"
            ),

            average_gor=Avg(
                "gor"
            ),

            latest_test_date=Max(
                "test_date"
            ),
        )

        # -----------------------------
        # Test History
        # -----------------------------

        tests = queryset.order_by(
            "test_date"
        )

        test_history = [
            {
                "date": test.test_date,
                "well": test.well.code,

                "oil_rate": float(
                    test.oil_rate or 0
                ),

                "gas_rate": float(
                    test.gas_rate or 0
                ),

                "water_rate": float(
                    test.water_rate or 0
                ),

                "wellhead_pressure": (
                    float(test.wellhead_pressure)
                    if test.wellhead_pressure is not None
                    else None
                ),

                "bottomhole_pressure": (
                    float(test.bottomhole_pressure)
                    if test.bottomhole_pressure is not None
                    else None
                ),

                "choke_size": (
                    float(test.choke_size)
                    if test.choke_size is not None
                    else None
                ),

                "water_cut": (
                    float(test.water_cut)
                    if test.water_cut is not None
                    else None
                ),

                "gor": (
                    float(test.gor)
                    if test.gor is not None
                    else None
                ),
            }
            for test in tests
        ]

        # -----------------------------
        # Pressure Analysis
        # -----------------------------

        pressure_history = [
            {
                "date": test.test_date,
                "well": test.well.code,

                "wellhead_pressure": (
                    float(test.wellhead_pressure)
                    if test.wellhead_pressure is not None
                    else None
                ),

                "bottomhole_pressure": (
                    float(test.bottomhole_pressure)
                    if test.bottomhole_pressure is not None
                    else None
                ),
            }
            for test in tests
        ]

        # -----------------------------
        # Production Rates
        # -----------------------------

        production_history = [
            {
                "date": test.test_date,
                "well": test.well.code,

                "oil_rate": float(
                    test.oil_rate or 0
                ),

                "gas_rate": float(
                    test.gas_rate or 0
                ),

                "water_rate": float(
                    test.water_rate or 0
                ),
            }
            for test in tests
        ]
        well_performance = (
            queryset
            .values(
                "well__code",
                "well__name",
            )
            .annotate(
                average_oil_rate=Avg("oil_rate"),
                average_gas_rate=Avg("gas_rate"),
                average_water_rate=Avg("water_rate"),
                test_count=Count("id"),
            )
            .order_by("-average_oil_rate")
        )
        declining_wells = []

        well_ids = (
            queryset
            .values_list("well_id", flat=True)
            .distinct()
        )

        for well_id in well_ids:
            tests = list(
                queryset
                .filter(well_id=well_id)
                .order_by("test_date")
            )

            if len(tests) < 2:
                continue

            first = float(tests[0].oil_rate or 0)
            latest = float(tests[-1].oil_rate or 0)

            if first <= 0:
                continue

            decline_percentage = (
                (first - latest) / first
            ) * 100

            if decline_percentage > 0:
                declining_wells.append({
                    "well": tests[-1].well.code,
                    "first_oil_rate": first,
                    "latest_oil_rate": latest,
                    "decline_percentage": round(
                        decline_percentage,
                        2,
                    ),
                    "first_date": tests[0].test_date,
                    "latest_date": tests[-1].test_date,
                })

        declining_wells.sort(
            key=lambda x: x["decline_percentage"],
            reverse=True,
        )
        return Response({
            "report": "Well Test Report",

            "filters": {
                "field": field_id,
                "well": well_id,
                "date_from": date_from,
                "date_to": date_to,
            },

            "summary": summary,

            "test_history": test_history,

            "pressure_history": pressure_history,

            "production_history": production_history,
            "well_performance": [
                {
                    "well": item["well__code"],
                    "name": item["well__name"],
                    "average_oil_rate": float(
                        item["average_oil_rate"] or 0
                    ),
                    "average_gas_rate": float(
                        item["average_gas_rate"] or 0
                    ),
                    "average_water_rate": float(
                        item["average_water_rate"] or 0
                    ),
                    "test_count": item["test_count"],
                }
                for item in well_performance
            ],
            "declining_wells": declining_wells,
        })


class WellTestExcelExportView(APIView):
    permission_classes = [ReportPermission]

    def get(self, request):
        queryset = scope_queryset_for_user(
            request.user,
            WellTest.objects.select_related(
                "well",
                "well__field",
            ),
            "well_tests",
        )

        # Filters
        field_id = request.query_params.get("field")
        well_id = request.query_params.get("well")
        date_from = request.query_params.get("date_from")
        date_to = request.query_params.get("date_to")

        if field_id:
            queryset = queryset.filter(
                well__field_id=field_id
            )

        if well_id:
            queryset = queryset.filter(
                well_id=well_id
            )

        if date_from:
            queryset = queryset.filter(
                test_date__gte=date_from
            )

        if date_to:
            queryset = queryset.filter(
                test_date__lte=date_to
            )

        queryset = queryset.order_by(
            "-test_date"
        )

        # Workbook
        workbook = Workbook()
        summary_sheet = workbook.active
        summary_sheet.title = "Summary"
        summary_sheet.append(["Metric", "Value", "Unit"])

        summary = queryset.aggregate(
            total_tests=Count("id"),
            average_oil_rate=Avg("oil_rate"),
            average_gas_rate=Avg("gas_rate"),
            average_water_rate=Avg("water_rate"),
            average_wellhead_pressure=Avg("wellhead_pressure"),
            average_bottomhole_pressure=Avg("bottomhole_pressure"),
            average_water_cut=Avg("water_cut"),
            average_gor=Avg("gor"),
        )

        summary_rows = [
            ["Total Tests", summary["total_tests"] or 0, ""],
            ["Average Oil Rate", summary["average_oil_rate"] or 0, "BOPD"],
            ["Average Gas Rate", summary["average_gas_rate"] or 0, "MSCFD"],
            ["Average Water Rate", summary["average_water_rate"] or 0, "BWPD"],
            ["Average Wellhead Pressure", summary["average_wellhead_pressure"] or 0, "psi"],
            ["Average Bottomhole Pressure", summary["average_bottomhole_pressure"] or 0, "psi"],
            ["Average Water Cut", summary["average_water_cut"] or 0, "%"],
            ["Average GOR", summary["average_gor"] or 0, "scf/STB"],
        ]

        for row in summary_rows:
            summary_sheet.append(row)

        for cell in summary_sheet[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        summary_sheet.column_dimensions["A"].width = 32
        summary_sheet.column_dimensions["B"].width = 18
        summary_sheet.column_dimensions["C"].width = 14

        worksheet = workbook.create_sheet("Well Tests")

        headers = [
            "Date",
            "Well",
            "Field",
            "Oil Rate (BOPD)",
            "Gas Rate (MSCFD)",
            "Water Rate (BWPD)",
            "Wellhead Pressure (psi)",
            "Bottomhole Pressure (psi)",
            "Choke Size (1/64 in)",
            "Water Cut (%)",
            "GOR (scf/STB)",
            "Remarks",
        ]

        worksheet.append(headers)

        # Header styling
        for cell in worksheet[1]:
            cell.font = Font(
                bold=True
            )

            cell.alignment = Alignment(
                horizontal="center"
            )

        # Data
        for test in queryset:
            worksheet.append([
                test.test_date,
                test.well.code,
                test.well.field.name
                if test.well.field
                else "",

                float(test.oil_rate or 0),

                float(test.gas_rate or 0),

                float(test.water_rate or 0),

                (
                    float(test.wellhead_pressure)
                    if test.wellhead_pressure is not None
                    else None
                ),

                (
                    float(test.bottomhole_pressure)
                    if test.bottomhole_pressure is not None
                    else None
                ),

                (
                    float(test.choke_size)
                    if test.choke_size is not None
                    else None
                ),

                (
                    float(test.water_cut)
                    if test.water_cut is not None
                    else None
                ),

                (
                    float(test.gor)
                    if test.gor is not None
                    else None
                ),

                test.remarks or "",
            ])

        # Column widths
        widths = [
            15,
            15,
            25,
            18,
            18,
            18,
            25,
            27,
            22,
            18,
            18,
            35,
        ]

        for index, width in enumerate(
            widths,
            start=1
        ):
            worksheet.column_dimensions[
                chr(64 + index)
            ].width = width

        # Freeze header
        worksheet.freeze_panes = "A2"

        # Response
        response = HttpResponse(
            content_type=(
                "application/vnd.openxmlformats-"
                "officedocument.spreadsheetml.sheet"
            )
        )

        response[
            "Content-Disposition"
        ] = (
            'attachment; '
            'filename="well_test_report.xlsx"'
        )

        workbook.save(response)

        return response


class WellTestPDFExportView(APIView):
    permission_classes = [ReportPermission]

    def get(self, request):
        queryset = scope_queryset_for_user(
            request.user,
            WellTest.objects.select_related(
                "well",
                "well__field",
            ),
            "well_tests",
        )

        field_id = request.query_params.get("field")
        well_id = request.query_params.get("well")
        date_from = request.query_params.get("date_from")
        date_to = request.query_params.get("date_to")

        if field_id:
            queryset = queryset.filter(
                well__field_id=field_id
            )

        if well_id:
            queryset = queryset.filter(
                well_id=well_id
            )

        if date_from:
            queryset = queryset.filter(
                test_date__gte=date_from
            )

        if date_to:
            queryset = queryset.filter(
                test_date__lte=date_to
            )

        queryset = queryset.order_by(
            "-test_date"
        )

        filters = {
            "field": field_id,
            "well": well_id,
            "date_from": date_from,
            "date_to": date_to,
        }
        well_performance = (
            queryset
            .values(
                "well__code",
                "well__name",
            )
            .annotate(
                average_oil_rate=Avg("oil_rate"),
                average_gas_rate=Avg("gas_rate"),
                average_water_rate=Avg("water_rate"),
                test_count=Count("id"),
            )
            .order_by("-average_oil_rate")
        )

        well_performance = [
            {
                "well": item["well__code"],
                "name": item["well__name"],
                "average_oil_rate": float(
                    item["average_oil_rate"] or 0
                ),
                "average_gas_rate": float(
                    item["average_gas_rate"] or 0
                ),
                "average_water_rate": float(
                    item["average_water_rate"] or 0
                ),
                "test_count": item["test_count"],
            }
            for item in well_performance
        ]
        return export_well_test_pdf(
            queryset,
            filters,
            well_performance,
        )


class MeasurementReportView(APIView):
    permission_classes = [ReportPermission]

    def get_queryset(self, request):
        queryset = scope_queryset_for_user(
            request.user,
            WellMeasurement.objects.select_related(
                "well",
                "well__field",
                "recorded_by",
                "downtime_reason",
            ),
            "measurements",
        )

        field_id = request.query_params.get("field")
        well_id = request.query_params.get("well")
        shift = request.query_params.get("shift")
        operating_status = request.query_params.get("operating_status")
        date_from = request.query_params.get("date_from")
        date_to = request.query_params.get("date_to")

        if field_id:
            queryset = queryset.filter(well__field_id=field_id)
        if well_id:
            queryset = queryset.filter(well_id=well_id)
        if shift:
            queryset = queryset.filter(shift=shift)
        if operating_status:
            queryset = queryset.filter(operating_status=operating_status)

        if date_from:
            try:
                parsed_from = datetime.strptime(date_from, "%Y-%m-%d").date()
            except ValueError as error:
                raise ValueError("Invalid date_from format. Use YYYY-MM-DD.") from error
            queryset = queryset.filter(measurement_date__gte=parsed_from)

        if date_to:
            try:
                parsed_to = datetime.strptime(date_to, "%Y-%m-%d").date()
            except ValueError as error:
                raise ValueError("Invalid date_to format. Use YYYY-MM-DD.") from error
            queryset = queryset.filter(measurement_date__lte=parsed_to)

        return queryset

    @staticmethod
    def _number(value):
        return float(value) if value is not None else None

    def build_report(self, request):
        queryset = self.get_queryset(request)
        averages = queryset.aggregate(
            average_wellhead_pressure=Avg("wellhead_pressure"),
            average_tubing_head_pressure=Avg("tubing_head_pressure"),
            average_casing_pressure=Avg("casing_pressure"),
            average_flowline_pressure=Avg("flowline_pressure"),
            average_wellhead_temperature=Avg("wellhead_temperature"),
            average_flowline_temperature=Avg("flowline_temperature"),
            average_choke_size=Avg("choke_size"),
            average_esp_frequency=Avg("esp_frequency"),
            average_motor_current=Avg("motor_current"),
        )

        temperature_values = [
            value for value in [
                averages["average_wellhead_temperature"],
                averages["average_flowline_temperature"],
            ] if value is not None
        ]

        summary = {
            "total_measurements": queryset.count(),
            **{
                key: self._number(value)
                for key, value in averages.items()
            },
            "average_temperature": self._number(
                sum(temperature_values) / len(temperature_values)
                if temperature_values else None
            ),
        }

        pressure_history = []
        temperature_history = []
        choke_history = []
        esp_history = []
        history = []

        for measurement in queryset.order_by("measurement_date", "well__code"):
            date = measurement.measurement_date.isoformat()
            base = {
                "date": date,
                "well": measurement.well.code,
            }
            pressure_history.append({
                **base,
                "wellhead_pressure": self._number(measurement.wellhead_pressure),
                "tubing_head_pressure": self._number(measurement.tubing_head_pressure),
                "casing_pressure": self._number(measurement.casing_pressure),
                "flowline_pressure": self._number(measurement.flowline_pressure),
            })
            temperature_history.append({
                **base,
                "wellhead_temperature": self._number(measurement.wellhead_temperature),
                "flowline_temperature": self._number(measurement.flowline_temperature),
            })
            choke_history.append({
                **base,
                "choke_size": self._number(measurement.choke_size),
            })
            esp_history.append({
                **base,
                "esp_frequency": self._number(measurement.esp_frequency),
                "motor_current": self._number(measurement.motor_current),
            })

            recorded_by = ""
            if measurement.recorded_by:
                recorded_by = (
                    measurement.recorded_by.get_full_name()
                    or measurement.recorded_by.username
                )
            history.append({
                "id": measurement.id,
                "date": date,
                "well": measurement.well.code,
                "well_name": measurement.well.name,
                "field": measurement.well.field.name,
                "shift": measurement.shift,
                "operating_status": measurement.operating_status,
                "recorded_by": recorded_by,
                "wellhead_pressure": self._number(measurement.wellhead_pressure),
                "tubing_head_pressure": self._number(measurement.tubing_head_pressure),
                "casing_pressure": self._number(measurement.casing_pressure),
                "flowline_pressure": self._number(measurement.flowline_pressure),
                "wellhead_temperature": self._number(measurement.wellhead_temperature),
                "flowline_temperature": self._number(measurement.flowline_temperature),
                "choke_size": self._number(measurement.choke_size),
                "esp_frequency": self._number(measurement.esp_frequency),
                "motor_current": self._number(measurement.motor_current),
                "water_cut": self._number(measurement.water_cut),
                "gor": self._number(measurement.gor),
                "bsw": self._number(measurement.bsw),
                "downtime_hours": self._number(measurement.downtime_hours),
                "downtime_reason": measurement.downtime_reason.name
                if measurement.downtime_reason else "",
                "remarks": measurement.remarks or "",
            })

        status_distribution = [
            {"status": item["operating_status"], "count": item["count"]}
            for item in queryset.values("operating_status")
            .annotate(count=Count("id"))
            .order_by("-count", "operating_status")
        ]
        shift_distribution = [
            {"shift": item["shift"], "count": item["count"]}
            for item in queryset.values("shift")
            .annotate(count=Count("id"))
            .order_by("-count", "shift")
        ]

        return {
            "report": "Measurement Report",
            "filters": {
                "field": request.query_params.get("field"),
                "well": request.query_params.get("well"),
                "shift": request.query_params.get("shift"),
                "operating_status": request.query_params.get("operating_status"),
                "date_from": request.query_params.get("date_from"),
                "date_to": request.query_params.get("date_to"),
            },
            "summary": summary,
            "pressure_history": pressure_history,
            "temperature_history": temperature_history,
            "choke_history": choke_history,
            "esp_history": esp_history,
            "status_distribution": status_distribution,
            "shift_distribution": shift_distribution,
            "history": history,
        }

    def get(self, request):
        try:
            return Response(self.build_report(request))
        except ValueError as error:
            return Response({"error": str(error)}, status=400)


class MeasurementReportExcelView(MeasurementReportView):
    def get(self, request):
        try:
            report_data = self.build_report(request)
        except ValueError as error:
            return Response({"error": str(error)}, status=400)

        workbook = create_measurement_workbook(report_data)
        response = HttpResponse(
            content_type=(
                "application/vnd.openxmlformats-"
                "officedocument.spreadsheetml.sheet"
            )
        )
        response["Content-Disposition"] = 'attachment; filename="measurement_report.xlsx"'
        workbook.save(response)
        return response


class MeasurementReportPDFView(MeasurementReportView):
    def get(self, request):
        try:
            report_data = self.build_report(request)
        except ValueError as error:
            return Response({"error": str(error)}, status=400)

        return FileResponse(
            generate_measurement_pdf(report_data),
            as_attachment=True,
            filename="measurement_report.pdf",
            content_type="application/pdf",
        )


class FieldReportView(APIView):
    permission_classes = [ReportPermission]

    def get_field_queryset(self, request):
        queryset = scope_queryset_for_user(
            request.user,
            Field.objects.select_related("operator"),
            "fields",
        )
        operator = request.query_params.get("operator")
        status = request.query_params.get("status")
        location = request.query_params.get("location")

        if operator:
            queryset = queryset.filter(operator_id=operator)
        if status:
            queryset = queryset.filter(status=status)
        if location:
            queryset = queryset.filter(city__iexact=location)

        return queryset

    def build_report(self, request):
        fields = self.get_field_queryset(request)
        field_ids = fields.values("id")

        date_from = request.query_params.get("date_from")
        date_to = request.query_params.get("date_to")
        production = Production.objects.filter(
            well__field_id__in=field_ids
        ).select_related("well", "well__field")

        if date_from:
            try:
                date_from_value = datetime.strptime(
                    date_from, "%Y-%m-%d"
                ).date()
            except ValueError as error:
                raise ValueError(
                    "Invalid date_from format. Use YYYY-MM-DD."
                ) from error
            production = production.filter(
                production_date__gte=date_from_value
            )

        if date_to:
            try:
                date_to_value = datetime.strptime(
                    date_to, "%Y-%m-%d"
                ).date()
            except ValueError as error:
                raise ValueError(
                    "Invalid date_to format. Use YYYY-MM-DD."
                ) from error
            production = production.filter(
                production_date__lte=date_to_value
            )

        wells = Well.objects.filter(field_id__in=field_ids)
        summary_production = production.aggregate(
            total_oil=Sum("oil_production"),
            total_gas=Sum("gas_production"),
            total_water=Sum("water_production"),
        )

        field_counts = fields.annotate(
            well_count=Count("wells"),
            producing_wells=Count(
                "wells",
                filter=Q(wells__status=Well.Status.PRODUCING),
            ),
            shut_in_wells=Count(
                "wells",
                filter=Q(wells__status=Well.Status.SHUT_IN),
            ),
            drilling_wells=Count(
                "wells",
                filter=Q(wells__status=Well.Status.DRILLING),
            ),
        )

        summary = {
            "total_fields": fields.count(),
            "active_fields": fields.filter(
                status=Field.Status.ACTIVE
            ).count(),
            "total_wells": wells.count(),
            "producing_wells": wells.filter(
                status=Well.Status.PRODUCING
            ).count(),
            "total_operators": fields.values(
                "operator_id"
            ).distinct().count(),
            "total_oil": float(summary_production["total_oil"] or 0),
            "total_gas": float(summary_production["total_gas"] or 0),
            "total_water": float(summary_production["total_water"] or 0),
        }

        status_distribution = [
            {"status": item["status"], "count": item["count"]}
            for item in fields.values("status")
            .annotate(count=Count("id"))
            .order_by("-count", "status")
        ]
        operator_distribution = [
            {
                "operator": item["operator__short_name"],
                "count": item["count"],
            }
            for item in fields.values(
                "operator__short_name"
            ).annotate(count=Count("id")).order_by("-count")
        ]
        location_distribution = [
            {
                "location": item["city"],
                "count": item["count"],
            }
            for item in fields.exclude(city__isnull=True).exclude(
                city=""
            ).values("city").annotate(
                count=Count("id")
            ).order_by("-count", "city")
        ]
        well_distribution = [
            {
                "field": item.name,
                "count": item.well_count,
                "producing_wells": item.producing_wells,
                "shut_in_wells": item.shut_in_wells,
                "drilling_wells": item.drilling_wells,
            }
            for item in field_counts.order_by("-well_count", "name")
        ]
        well_type_distribution = [
            {
                "well_type": item["well_type"],
                "count": item["count"],
            }
            for item in wells.values("well_type")
            .annotate(count=Count("id"))
            .order_by("-count", "well_type")
        ]

        production_by_field = [
            {
                "field": item["well__field__name"],
                "oil": float(item["oil"] or 0),
                "gas": float(item["gas"] or 0),
                "water": float(item["water"] or 0),
            }
            for item in production.values("well__field__name")
            .annotate(
                oil=Sum("oil_production"),
                gas=Sum("gas_production"),
                water=Sum("water_production"),
            ).order_by("well__field__name")
        ]

        history = [
            {
                "id": field.id,
                "name": field.name,
                "code": field.code,
                "operator": field.operator.short_name,
                "country": field.country or "",
                "state": field.state or "",
                "city": field.city or "",
                "latitude": float(field.latitude)
                if field.latitude is not None else None,
                "longitude": float(field.longitude)
                if field.longitude is not None else None,
                "status": field.status,
                "well_count": field.well_count,
                "producing_wells": field.producing_wells,
                "shut_in_wells": field.shut_in_wells,
                "drilling_wells": field.drilling_wells,
            }
            for field in field_counts.order_by("name")
        ]

        return {
            "report": "Field Report",
            "filters": {
                "operator": request.query_params.get("operator"),
                "status": request.query_params.get("status"),
                "location": request.query_params.get("location"),
                "date_from": date_from,
                "date_to": date_to,
            },
            "summary": summary,
            "status_distribution": status_distribution,
            "operator_distribution": operator_distribution,
            "location_distribution": location_distribution,
            "well_distribution": well_distribution,
            "well_type_distribution": well_type_distribution,
            "production_by_field": production_by_field,
            "history": history,
        }

    def get(self, request):
        try:
            return Response(self.build_report(request))
        except ValueError as error:
            return Response({"error": str(error)}, status=400)


class FieldReportExcelView(FieldReportView):
    def get(self, request):
        try:
            report_data = self.build_report(request)
        except ValueError as error:
            return Response({"error": str(error)}, status=400)

        workbook = create_field_workbook(report_data)
        response = HttpResponse(
            content_type=(
                "application/vnd.openxmlformats-"
                "officedocument.spreadsheetml.sheet"
            )
        )
        response[
            "Content-Disposition"
        ] = 'attachment; filename="field_report.xlsx"'
        workbook.save(response)
        return response


class FieldReportPDFView(FieldReportView):
    def get(self, request):
        try:
            report_data = self.build_report(request)
        except ValueError as error:
            return Response({"error": str(error)}, status=400)

        return FileResponse(
            generate_field_pdf(report_data),
            as_attachment=True,
            filename="field_report.pdf",
            content_type="application/pdf",
        )


class MaintenanceReportView(APIView):
    permission_classes = [ReportPermission]

    def get_queryset(self, request):
        queryset = Maintenance.objects.select_related(
            "well",
            "well__field",
            "service_company",
            "assigned_to",
        )

        field_id = request.query_params.get("field")
        well_id = request.query_params.get("well")
        maintenance_type = request.query_params.get("maintenance_type")
        status = request.query_params.get("status")
        date_from = request.query_params.get("date_from")
        date_to = request.query_params.get("date_to")

        if field_id:
            queryset = queryset.filter(well__field_id=field_id)

        if well_id:
            queryset = queryset.filter(well_id=well_id)

        if maintenance_type:
            queryset = queryset.filter(maintenance_type=maintenance_type)

        if status:
            queryset = queryset.filter(status=status)

        if date_from:
            try:
                parsed_date_from = datetime.strptime(
                    date_from,
                    "%Y-%m-%d",
                ).date()
            except ValueError as error:
                raise ValueError(
                    "Invalid date_from format. Use YYYY-MM-DD."
                ) from error

            queryset = queryset.filter(start_date__gte=parsed_date_from)

        if date_to:
            try:
                parsed_date_to = datetime.strptime(
                    date_to,
                    "%Y-%m-%d",
                ).date()
            except ValueError as error:
                raise ValueError(
                    "Invalid date_to format. Use YYYY-MM-DD."
                ) from error

            queryset = queryset.filter(start_date__lte=parsed_date_to)

        return queryset

    @staticmethod
    def _date_value(value):
        return value.isoformat() if value else None

    @staticmethod
    def _time_value(value):
        return value.strftime("%H:%M") if value else None

    def build_report(self, request):
        queryset = self.get_queryset(request)

        summary_counts = queryset.aggregate(
            total_maintenance=Count("id"),
            completed=Count(
                "id",
                filter=Q(status=Maintenance.Status.COMPLETED),
            ),
            in_progress=Count(
                "id",
                filter=Q(status=Maintenance.Status.IN_PROGRESS),
            ),
            planned=Count(
                "id",
                filter=Q(status=Maintenance.Status.PLANNED),
            ),
            cancelled=Count(
                "id",
                filter=Q(status=Maintenance.Status.CANCELLED),
            ),
        )

        type_distribution = [
            {
                "type": item["maintenance_type"] or "Unknown",
                "count": item["count"],
            }
            for item in queryset.values("maintenance_type")
            .annotate(count=Count("id"))
            .order_by("-count", "maintenance_type")
        ]

        status_distribution = [
            {
                "status": item["status"] or "Unknown",
                "count": item["count"],
            }
            for item in queryset.values("status")
            .annotate(count=Count("id"))
            .order_by("-count", "status")
        ]

        well_distribution = [
            {
                "well": item["well__code"],
                "well_name": item["well__name"],
                "count": item["count"],
            }
            for item in queryset.values(
                "well__code",
                "well__name",
            )
            .annotate(count=Count("id"))
            .order_by("-count", "well__code")
        ]

        field_distribution = [
            {
                "field": item["well__field__name"],
                "count": item["count"],
            }
            for item in queryset.values("well__field__name")
            .annotate(count=Count("id"))
            .order_by("-count", "well__field__name")
        ]

        timeline_queryset = queryset.annotate(
            month=TruncMonth("start_date")
        ).values("month").annotate(
            count=Count("id"),
        ).order_by("month")

        timeline = [
            {
                "date": item["month"].strftime("%Y-%m"),
                "count": item["count"],
            }
            for item in timeline_queryset
        ]

        comparison_queryset = queryset.annotate(
            month=TruncMonth("start_date")
        ).values("month").annotate(
            completed=Count(
                "id",
                filter=Q(status=Maintenance.Status.COMPLETED),
            ),
            pending=Count(
                "id",
                filter=~Q(status=Maintenance.Status.COMPLETED),
            ),
        ).order_by("month")

        completed_vs_pending = [
            {
                "date": item["month"].strftime("%Y-%m"),
                "completed": item["completed"],
                "pending": item["pending"],
            }
            for item in comparison_queryset
        ]

        history = []
        for maintenance in queryset.order_by(
            "-start_date",
            "-start_time",
        ):
            assigned_to = ""
            if maintenance.assigned_to:
                assigned_to = (
                    maintenance.assigned_to.get_full_name()
                    or maintenance.assigned_to.username
                )

            history.append({
                "id": maintenance.id,
                "well": maintenance.well.code,
                "well_name": maintenance.well.name,
                "field": maintenance.well.field.name,
                "maintenance_type": maintenance.maintenance_type,
                "title": maintenance.title,
                "description": maintenance.description or "",
                "service_company": (
                    maintenance.service_company.short_name
                    if maintenance.service_company
                    else ""
                ),
                "assigned_to": assigned_to,
                "start_date": self._date_value(maintenance.start_date),
                "start_time": self._time_value(maintenance.start_time),
                "end_date": self._date_value(maintenance.end_date),
                "end_time": self._time_value(maintenance.end_time),
                "estimated_cost": (
                    float(maintenance.estimated_cost)
                    if maintenance.estimated_cost is not None
                    else None
                ),
                "actual_cost": (
                    float(maintenance.actual_cost)
                    if maintenance.actual_cost is not None
                    else None
                ),
                "status": maintenance.status,
                "remarks": maintenance.remarks or "",
            })

        return {
            "report": "Maintenance Report",
            "filters": {
                "field": request.query_params.get("field"),
                "well": request.query_params.get("well"),
                "maintenance_type": request.query_params.get(
                    "maintenance_type"
                ),
                "status": request.query_params.get("status"),
                "date_from": request.query_params.get("date_from"),
                "date_to": request.query_params.get("date_to"),
            },
            "summary": summary_counts,
            "type_distribution": type_distribution,
            "status_distribution": status_distribution,
            "well_distribution": well_distribution,
            "field_distribution": field_distribution,
            "timeline": timeline,
            "completed_vs_pending": completed_vs_pending,
            "history": history,
        }

    def get(self, request):
        try:
            return Response(self.build_report(request))
        except ValueError as error:
            return Response({"error": str(error)}, status=400)


class MaintenanceReportExcelView(MaintenanceReportView):
    def get(self, request):
        try:
            report_data = self.build_report(request)
        except ValueError as error:
            return Response({"error": str(error)}, status=400)

        workbook = create_maintenance_workbook(report_data)
        response = HttpResponse(
            content_type=(
                "application/vnd.openxmlformats-"
                "officedocument.spreadsheetml.sheet"
            )
        )
        response[
            "Content-Disposition"
        ] = 'attachment; filename="maintenance_report.xlsx"'
        workbook.save(response)
        return response


class MaintenanceReportPDFView(MaintenanceReportView):
    def get(self, request):
        try:
            report_data = self.build_report(request)
        except ValueError as error:
            return Response({"error": str(error)}, status=400)

        return FileResponse(
            generate_maintenance_pdf(report_data),
            as_attachment=True,
            filename="maintenance_report.pdf",
            content_type="application/pdf",
        )


class WellReportView(APIView):
    permission_classes = [ReportPermission]

    def get_queryset(self, request):
        queryset = scope_queryset_for_user(
            request.user,
            Well.objects.select_related(
                "field",
                "operator",
            ),
            "wells",
        )

        field_id = request.query_params.get("field")
        operator_id = request.query_params.get("operator")
        well_type = request.query_params.get("well_type")
        status = request.query_params.get("status")
        artificial_lift = request.query_params.get("artificial_lift")
        is_active = request.query_params.get("is_active")

        if field_id:
            queryset = queryset.filter(field_id=field_id)

        if operator_id:
            queryset = queryset.filter(operator_id=operator_id)

        if well_type:
            queryset = queryset.filter(well_type=well_type)

        if status:
            queryset = queryset.filter(status=status)

        if artificial_lift:
            queryset = queryset.filter(
                artificial_lift=artificial_lift
            )

        if is_active is not None:
            if is_active.lower() == "true":
                queryset = queryset.filter(is_active=True)
            elif is_active.lower() == "false":
                queryset = queryset.filter(is_active=False)

        return queryset

    def build_report(self, request):
        queryset = self.get_queryset(request)

        total_wells = queryset.count()
        active_wells = queryset.filter(is_active=True).count()
        producing_wells = queryset.filter(
            status=Well.Status.PRODUCING
        ).count()
        drilling_wells = queryset.filter(
            status=Well.Status.DRILLING
        ).count()
        shut_in_wells = queryset.filter(
            status=Well.Status.SHUT_IN
        ).count()
        abandoned_wells = queryset.filter(
            status=Well.Status.ABANDONED
        ).count()

        averages = queryset.aggregate(
            average_total_depth=Avg("total_depth"),
            average_true_vertical_depth=Avg(
                "true_vertical_depth"
            ),
        )

        summary = {
            "total_wells": total_wells,
            "active_wells": active_wells,
            "producing_wells": producing_wells,
            "drilling_wells": drilling_wells,
            "shut_in_wells": shut_in_wells,
            "abandoned_wells": abandoned_wells,
            "average_total_depth": float(
                averages["average_total_depth"] or 0
            ),
            "average_true_vertical_depth": float(
                averages["average_true_vertical_depth"] or 0
            ),
        }

        well_type_distribution = [
            {
                "type": item["well_type"] or "Unknown",
                "count": item["count"],
            }
            for item in queryset.values("well_type").annotate(
                count=Count("id")
            ).order_by("-count")
        ]

        status_distribution = [
            {
                "status": item["status"] or "Unknown",
                "count": item["count"],
            }
            for item in queryset.values("status").annotate(
                count=Count("id")
            ).order_by("-count")
        ]

        field_distribution = [
            {
                "field": item["field__name"],
                "count": item["count"],
            }
            for item in queryset.values(
                "field__name"
            ).annotate(count=Count("id")).order_by("-count")
        ]

        operator_distribution = [
            {
                "operator": item["operator__short_name"],
                "count": item["count"],
            }
            for item in queryset.values(
                "operator__short_name"
            ).annotate(count=Count("id")).order_by("-count")
        ]

        artificial_lift_distribution = [
            {
                "artificial_lift": item["artificial_lift"]
                or "Unknown",
                "count": item["count"],
            }
            for item in queryset.values(
                "artificial_lift"
            ).annotate(count=Count("id")).order_by("-count")
        ]

        reservoir_distribution = [
            {
                "reservoir": item["reservoir"] or "Unknown",
                "count": item["count"],
            }
            for item in queryset.values("reservoir").annotate(
                count=Count("id")
            ).order_by("-count")
        ]

        formation_distribution = [
            {
                "formation": item["formation"] or "Unknown",
                "count": item["count"],
            }
            for item in queryset.values("formation").annotate(
                count=Count("id")
            ).order_by("-count")
        ]

        spud_data = queryset.filter(
            spud_date__isnull=False
        ).annotate(
            month=TruncMonth("spud_date")
        ).values("month").annotate(
            count=Count("id")
        ).order_by("month")
        completion_data = queryset.filter(
            completion_date__isnull=False
        ).annotate(
            month=TruncMonth("completion_date")
        ).values("month").annotate(
            count=Count("id")
        ).order_by("month")
        production_data = queryset.filter(
            first_production_date__isnull=False
        ).annotate(
            month=TruncMonth("first_production_date")
        ).values("month").annotate(
            count=Count("id")
        ).order_by("month")

        timeline_map = {}
        for item in spud_data:
            label = item["month"].strftime("%Y-%m")
            timeline_map.setdefault(label, {})["spud"] = item["count"]
        for item in completion_data:
            label = item["month"].strftime("%Y-%m")
            timeline_map.setdefault(label, {})["completed"] = item["count"]
        for item in production_data:
            label = item["month"].strftime("%Y-%m")
            timeline_map.setdefault(label, {})["first_production"] = item["count"]

        timeline = [
            {
                "month": month,
                "spud": timeline_map[month].get("spud", 0),
                "completed": timeline_map[month].get("completed", 0),
                "first_production": timeline_map[month].get(
                    "first_production", 0
                ),
            }
            for month in sorted(timeline_map.keys())
        ]

        locations = []
        for well in queryset.filter(
            latitude__isnull=False,
            longitude__isnull=False,
        ):
            locations.append(
                {
                    "code": well.code,
                    "name": well.name,
                    "field": well.field.name
                    if well.field
                    else "",
                    "operator": well.operator.short_name
                    if well.operator
                    else "",
                    "latitude": float(well.latitude),
                    "longitude": float(well.longitude),
                }
            )

        history = []
        for well in queryset.order_by("code"):
            history.append(
                {
                    "code": well.code,
                    "name": well.name,
                    "field": well.field.name if well.field else "",
                    "operator": well.operator.short_name
                    if well.operator
                    else "",
                    "well_type": well.well_type,
                    "status": well.status,
                    "spud_date": well.spud_date.isoformat()
                    if well.spud_date
                    else None,
                    "completion_date": well.completion_date.isoformat()
                    if well.completion_date
                    else None,
                    "first_production_date": well.first_production_date.isoformat()
                    if well.first_production_date
                    else None,
                    "total_depth": float(well.total_depth)
                    if well.total_depth is not None
                    else None,
                    "true_vertical_depth": float(
                        well.true_vertical_depth
                    )
                    if well.true_vertical_depth is not None
                    else None,
                    "tubing_size": well.tubing_size or "",
                    "casing_size": well.casing_size or "",
                    "artificial_lift": well.artificial_lift or "",
                    "reservoir": well.reservoir or "",
                    "formation": well.formation or "",
                    "latitude": float(well.latitude)
                    if well.latitude is not None
                    else None,
                    "longitude": float(well.longitude)
                    if well.longitude is not None
                    else None,
                    "is_active": well.is_active,
                }
            )

        return {
            "report": "Well Report",
            "filters": {
                "field": request.query_params.get("field"),
                "operator": request.query_params.get("operator"),
                "well_type": request.query_params.get("well_type"),
                "status": request.query_params.get("status"),
                "artificial_lift": request.query_params.get(
                    "artificial_lift"
                ),
                "is_active": request.query_params.get("is_active"),
            },
            "summary": summary,
            "well_type_distribution": well_type_distribution,
            "status_distribution": status_distribution,
            "field_distribution": field_distribution,
            "operator_distribution": operator_distribution,
            "artificial_lift_distribution": artificial_lift_distribution,
            "reservoir_distribution": reservoir_distribution,
            "formation_distribution": formation_distribution,
            "timeline": timeline,
            "locations": locations,
            "history": history,
        }

    def get(self, request):
        report_data = self.build_report(request)
        return Response(report_data)


class WellReportExcelView(WellReportView):
    def get(self, request):
        report_data = self.build_report(request)
        workbook = create_well_workbook(report_data)
        response = HttpResponse(
            content_type=(
                "application/vnd.openxmlformats-"
                "officedocument.spreadsheetml.sheet"
            )
        )
        response[
            "Content-Disposition"
        ] = 'attachment; filename="well_report.xlsx"'
        workbook.save(response)
        return response


class WellReportPDFView(WellReportView):
    def get(self, request):
        report_data = self.build_report(request)
        pdf_buffer = generate_well_pdf(report_data)
        return FileResponse(
            pdf_buffer,
            as_attachment=True,
            filename="well_report.pdf",
            content_type="application/pdf",
        )


class InterventionReportView(APIView):
    permission_classes = [ReportPermission]

    def get(self, request):
        queryset = scope_queryset_for_user(
            request.user,
            WellIntervention.objects.select_related(
                "well",
                "well__field",
                "service_company",
                "supervisor",
            ),
            "interventions",
        )

        # -----------------------------
        # Filters
        # -----------------------------

        field_id = request.query_params.get("field")
        well_id = request.query_params.get("well")
        intervention_type = request.query_params.get(
            "intervention_type"
        )
        status = request.query_params.get("status")
        date_from = request.query_params.get("date_from")
        date_to = request.query_params.get("date_to")

        if field_id:
            queryset = queryset.filter(
                well__field_id=field_id
            )

        if well_id:
            queryset = queryset.filter(
                well_id=well_id
            )

        if intervention_type:
            queryset = queryset.filter(
                intervention_type=intervention_type
            )

        if status:
            queryset = queryset.filter(
                status=status
            )

        if date_from:
            queryset = queryset.filter(
                start_date__gte=date_from
            )

        if date_to:
            queryset = queryset.filter(
                start_date__lte=date_to
            )

        queryset = queryset.order_by(
            "-start_date",
            "-start_time",
        )

        # -----------------------------
        # Executive Summary
        # -----------------------------

        total_interventions = queryset.count()

        completed = queryset.filter(
            status=WellIntervention.Status.COMPLETED
        ).count()

        in_progress = queryset.filter(
            status=WellIntervention.Status.IN_PROGRESS
        ).count()

        planned = queryset.filter(
            status=WellIntervention.Status.PLANNED
        ).count()

        cancelled = queryset.filter(
            status=WellIntervention.Status.CANCELLED
        ).count()

        # -----------------------------
        # History
        # -----------------------------

        history = []

        for intervention in queryset:
            history.append({
                "id": intervention.id,

                "well": intervention.well.code,

                "well_name": intervention.well.name,

                "field": (
                    intervention.well.field.name
                    if intervention.well.field
                    else None
                ),

                "intervention_type": (
                    intervention.intervention_type
                ),

                "title": intervention.title,

                "service_company": (
                    intervention.service_company.short_name
                    if intervention.service_company
                    else None
                ),

                "supervisor": (
                    intervention.supervisor.username
                    if intervention.supervisor
                    else None
                ),

                "start_date": (
                    intervention.start_date
                ),

                "start_time": (
                    intervention.start_time
                ),

                "end_date": (
                    intervention.end_date
                ),

                "end_time": (
                    intervention.end_time
                ),

                "status": intervention.status,

                "description": (
                    intervention.description
                ),

                "remarks": intervention.remarks,
            })
        # -----------------------------
        # Well Performance
        # -----------------------------

        well_performance = (
            queryset
            .values(
                "well__code",
                "well__name",
            )
            .annotate(
                intervention_count=Count("id")
            )
            .order_by("-intervention_count")
        )

        well_performance = [
            {
                "well": item["well__code"],
                "name": item["well__name"],
                "intervention_count": item[
                    "intervention_count"
                ],
            }
            for item in well_performance
        ]
        return Response({
            "report": "Intervention Report",

            "filters": {
                "field": field_id,
                "well": well_id,
                "intervention_type": (
                    intervention_type
                ),
                "status": status,
                "date_from": date_from,
                "date_to": date_to,
            },

            "summary": {
                "total_interventions":
                    total_interventions,

                "completed": completed,

                "in_progress": in_progress,

                "planned": planned,

                "cancelled": cancelled,
            },

            "history": history,
            "well_performance": well_performance,
        })

class InterventionReportExcelView(APIView):
    permission_classes = [ReportPermission]

    def get(self, request):

        queryset = scope_queryset_for_user(
            request.user,
            WellIntervention.objects.select_related(
                "well",
                "well__field",
                "service_company",
                "supervisor",
            ),
            "interventions",
        )

        # -----------------------------
        # Filters
        # -----------------------------

        field_id = request.query_params.get("field")
        well_id = request.query_params.get("well")
        intervention_type = request.query_params.get(
            "intervention_type"
        )
        status = request.query_params.get("status")
        date_from = request.query_params.get("date_from")
        date_to = request.query_params.get("date_to")

        if field_id:
            queryset = queryset.filter(
                well__field_id=field_id
            )

        if well_id:
            queryset = queryset.filter(
                well_id=well_id
            )

        if intervention_type:
            queryset = queryset.filter(
                intervention_type=intervention_type
            )

        if status:
            queryset = queryset.filter(
                status=status
            )

        if date_from:
            queryset = queryset.filter(
                start_date__gte=date_from
            )

        if date_to:
            queryset = queryset.filter(
                start_date__lte=date_to
            )

        queryset = queryset.order_by(
            "-start_date",
            "-start_time",
        )

        # -----------------------------
        # Workbook
        # -----------------------------

        workbook = Workbook()

        sheet = workbook.active
        sheet.title = "Interventions"

        headers = [
            "Well",
            "Well Name",
            "Field",
            "Intervention Type",
            "Title",
            "Service Company",
            "Supervisor",
            "Start Date",
            "Start Time",
            "End Date",
            "End Time",
            "Status",
            "Description",
            "Remarks",
        ]

        sheet.append(headers)

        # Header style
        for cell in sheet[1]:
            cell.font = Font(
                bold=True,
                color="FFFFFF",
            )

            cell.fill = PatternFill(
                "solid",
                fgColor="2563EB",
            )

            cell.alignment = Alignment(
                horizontal="center"
            )

        # -----------------------------
        # Data
        # -----------------------------

        for intervention in queryset:

            sheet.append([
                intervention.well.code,

                intervention.well.name,

                (
                    intervention.well.field.name
                    if intervention.well.field
                    else ""
                ),

                intervention.intervention_type,

                intervention.title,

                (
                    intervention.service_company.short_name
                    if intervention.service_company
                    else ""
                ),

                (
                    intervention.supervisor.username
                    if intervention.supervisor
                    else ""
                ),

                intervention.start_date,

                intervention.start_time,

                intervention.end_date,

                intervention.end_time,

                intervention.status,

                intervention.description or "",

                intervention.remarks or "",
            ])

        # -----------------------------
        # Column widths
        # -----------------------------

        widths = {
            "A": 15,
            "B": 30,
            "C": 25,
            "D": 25,
            "E": 35,
            "F": 20,
            "G": 20,
            "H": 15,
            "I": 12,
            "J": 15,
            "K": 12,
            "L": 15,
            "M": 50,
            "N": 50,
        }

        for column, width in widths.items():
            sheet.column_dimensions[
                column
            ].width = width

        # -----------------------------
        # Response
        # -----------------------------

        response = HttpResponse(
            content_type=(
                "application/vnd.openxmlformats-"
                "officedocument.spreadsheetml.sheet"
            )
        )

        response[
            "Content-Disposition"
        ] = (
            'attachment; '
            'filename="intervention_report.xlsx"'
        )

        workbook.save(response)

        return response

class InterventionReportPDFView(APIView):
    permission_classes = [ReportPermission]

    def get(self, request):

        queryset = scope_queryset_for_user(
            request.user,
            WellIntervention.objects.select_related(
                "well",
                "well__field",
                "service_company",
                "supervisor",
            ),
            "interventions",
        )

        # -----------------------------
        # Filters
        # -----------------------------

        field_id = request.query_params.get(
            "field"
        )

        well_id = request.query_params.get(
            "well"
        )

        intervention_type = (
            request.query_params.get(
                "intervention_type"
            )
        )

        status = request.query_params.get(
            "status"
        )

        date_from = request.query_params.get(
            "date_from"
        )

        date_to = request.query_params.get(
            "date_to"
        )

        # -----------------------------
        # Apply filters
        # -----------------------------

        if field_id:
            queryset = queryset.filter(
                well__field_id=field_id
            )

        if well_id:
            queryset = queryset.filter(
                well_id=well_id
            )

        if intervention_type:
            queryset = queryset.filter(
                intervention_type=intervention_type
            )

        if status:
            queryset = queryset.filter(
                status=status
            )

        if date_from:
            queryset = queryset.filter(
                start_date__gte=date_from
            )

        if date_to:
            queryset = queryset.filter(
                start_date__lte=date_to
            )

        queryset = queryset.order_by(
            "-start_date",
            "-start_time",
        )

        # -----------------------------
        # Filters object
        # -----------------------------

        filters = {
            "field": field_id,
            "well": well_id,
            "intervention_type":
                intervention_type,
            "status": status,
            "date_from": date_from,
            "date_to": date_to,
        }

        # -----------------------------
        # Generate PDF
        # -----------------------------

        pdf_buffer = (
            generate_intervention_pdf(
                queryset,
                filters,
            )
        )

        response = HttpResponse(
            pdf_buffer.getvalue(),
            content_type="application/pdf",
        )

        response[
            "Content-Disposition"
        ] = (
            'attachment; '
            'filename="intervention_report.pdf"'
        )

        return response