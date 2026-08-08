from openpyxl import Workbook
from openpyxl.styles import Alignment, Font


def _style_sheet(sheet, widths):
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for index, width in enumerate(widths, start=1):
        column = sheet.cell(row=1, column=index).column_letter
        sheet.column_dimensions[column].width = width
    sheet.freeze_panes = "A2"


def create_measurement_workbook(report_data):
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Summary"
    summary_sheet.append(["Metric", "Value", "Unit"])

    summary = report_data.get("summary", {})
    for key, label, unit in [
        ("total_measurements", "Total Measurements", ""),
        ("average_wellhead_pressure", "Average Wellhead Pressure", "psi"),
        ("average_tubing_head_pressure", "Average Tubing Head Pressure", "psi"),
        ("average_bottomhole_pressure", "Average Bottomhole Pressure", "psi"),
        ("average_casing_pressure", "Average Casing Pressure", "psi"),
        ("average_flowline_pressure", "Average Flowline Pressure", "psi"),
        ("average_wellhead_temperature", "Average Wellhead Temperature", "°C"),
        ("average_flowline_temperature", "Average Flowline Temperature", "°C"),
        ("average_choke_size", "Average Choke Size", "1/64 in"),
        ("average_esp_frequency", "Average ESP Frequency", "Hz"),
        ("average_motor_current", "Average Motor Current", "A"),
    ]:
        summary_sheet.append([label, summary.get(key, 0), unit])
    _style_sheet(summary_sheet, [38, 20, 14])

    history_sheet = workbook.create_sheet("Measurement History")
    headers = [
        "ID", "Date", "Well", "Well Name", "Field", "Shift", "Operating Status",
        "Recorded By", "Wellhead Pressure", "Tubing Head Pressure", "Casing Pressure",
        "Flowline Pressure", "Wellhead Temperature", "Flowline Temperature", "Choke Size",
        "ESP Frequency", "Motor Current", "Water Cut", "GOR", "BSW", "Downtime Hours",
        "Downtime Reason", "Remarks",
    ]
    history_sheet.append(headers)
    for item in report_data.get("history", []):
        history_sheet.append([
            item.get("id", ""), item.get("date", ""), item.get("well", ""),
            item.get("well_name", ""), item.get("field", ""), item.get("shift", ""),
            item.get("operating_status", ""), item.get("recorded_by", ""),
            item.get("wellhead_pressure", ""), item.get("tubing_head_pressure", ""),
            item.get("casing_pressure", ""), item.get("flowline_pressure", ""),
            item.get("wellhead_temperature", ""), item.get("flowline_temperature", ""),
            item.get("choke_size", ""), item.get("esp_frequency", ""),
            item.get("motor_current", ""), item.get("water_cut", ""),
            item.get("gor", ""), item.get("bsw", ""), item.get("downtime_hours", ""),
            item.get("downtime_reason", ""), item.get("remarks", ""),
        ])
    _style_sheet(history_sheet, [8, 15, 16, 28, 24, 12, 20, 24, 18, 20, 18, 18, 22, 22, 14, 16, 16, 14, 14, 14, 16, 24, 36])

    for title, key, label in [
        ("Pressure Analysis", "pressure_history", "Date"),
        ("Temperature", "temperature_history", "Date"),
        ("Choke", "choke_history", "Date"),
        ("ESP Performance", "esp_history", "Date"),
    ]:
        rows = report_data.get(key, [])
        if not rows:
            continue
        sheet = workbook.create_sheet(title)
        columns = list(rows[0].keys())
        sheet.append(columns)
        for row in rows:
            sheet.append([row.get(column, "") for column in columns])
        _style_sheet(sheet, [22] + [20] * (len(columns) - 1))

    for title, key, label in [
        ("Operating Status", "status_distribution", "Status"),
        ("Shift Analysis", "shift_distribution", "Shift"),
    ]:
        rows = report_data.get(key, [])
        sheet = workbook.create_sheet(title)
        sheet.append([label, "Count"])
        for row in rows:
            sheet.append([row.get(label.lower().replace(" ", "_"), ""), row.get("count", 0)])
        _style_sheet(sheet, [24, 16])

    return workbook
