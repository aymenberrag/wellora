from openpyxl import Workbook
from openpyxl.styles import Alignment, Font


CONTENT_TYPE = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)


def _style_sheet(sheet, widths):
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for index, width in enumerate(widths, start=1):
        column = sheet.cell(row=1, column=index).column_letter
        sheet.column_dimensions[column].width = width

    sheet.freeze_panes = "A2"


def _add_distribution_sheet(workbook, title, rows, first_header):
    sheet = workbook.create_sheet(title)
    sheet.append([first_header, "Count"])

    for row in rows:
        sheet.append([
            row.get(first_header.lower().replace(" ", "_"), ""),
            row.get("count", 0),
        ])

    _style_sheet(sheet, [32, 14])
    return sheet


def create_maintenance_workbook(report_data):
    workbook = Workbook()

    summary_sheet = workbook.active
    summary_sheet.title = "Summary"
    summary_sheet.append(["Metric", "Value"])

    summary = report_data.get("summary", {})
    for row in [
        ["Total Maintenance", summary.get("total_maintenance", 0)],
        ["Completed", summary.get("completed", 0)],
        ["In Progress", summary.get("in_progress", 0)],
        ["Planned", summary.get("planned", 0)],
        ["Cancelled", summary.get("cancelled", 0)],
    ]:
        summary_sheet.append(row)

    _style_sheet(summary_sheet, [32, 18])

    history_sheet = workbook.create_sheet("Maintenance History")
    history_headers = [
        "ID",
        "Well",
        "Well Name",
        "Field",
        "Maintenance Type",
        "Title",
        "Description",
        "Service Company",
        "Assigned To",
        "Start Date",
        "Start Time",
        "End Date",
        "End Time",
        "Estimated Cost",
        "Actual Cost",
        "Status",
        "Remarks",
    ]
    history_sheet.append(history_headers)

    for item in report_data.get("history", []):
        history_sheet.append([
            item.get("id", ""),
            item.get("well", ""),
            item.get("well_name", ""),
            item.get("field", ""),
            item.get("maintenance_type", ""),
            item.get("title", ""),
            item.get("description", ""),
            item.get("service_company", ""),
            item.get("assigned_to", ""),
            item.get("start_date", ""),
            item.get("start_time", ""),
            item.get("end_date", ""),
            item.get("end_time", ""),
            item.get("estimated_cost", ""),
            item.get("actual_cost", ""),
            item.get("status", ""),
            item.get("remarks", ""),
        ])

    _style_sheet(
        history_sheet,
        [8, 16, 28, 25, 20, 28, 40, 24, 24, 15, 12, 15, 12, 18, 18, 16, 40],
    )

    _add_distribution_sheet(
        workbook,
        "By Type",
        report_data.get("type_distribution", []),
        "Type",
    )
    _add_distribution_sheet(
        workbook,
        "By Status",
        report_data.get("status_distribution", []),
        "Status",
    )
    _add_distribution_sheet(
        workbook,
        "By Well",
        report_data.get("well_distribution", []),
        "Well",
    )
    _add_distribution_sheet(
        workbook,
        "By Field",
        report_data.get("field_distribution", []),
        "Field",
    )

    return workbook
