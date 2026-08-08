from openpyxl import Workbook
from openpyxl.styles import Alignment, Font


def create_well_workbook(report_data):
    workbook = Workbook()

    summary_sheet = workbook.active
    summary_sheet.title = "Summary"

    summary_sheet.append(["Metric", "Value"])
    summary_sheet["A1"].font = Font(bold=True)
    summary_sheet["B1"].font = Font(bold=True)

    summary = report_data.get("summary", {})
    rows = [
        ["Total Wells", summary.get("total_wells", 0)],
        ["Active Wells", summary.get("active_wells", 0)],
        ["Producing Wells", summary.get("producing_wells", 0)],
        ["Drilling Wells", summary.get("drilling_wells", 0)],
        ["Shut In Wells", summary.get("shut_in_wells", 0)],
        ["Abandoned Wells", summary.get("abandoned_wells", 0)],
        ["Average Total Depth", summary.get("average_total_depth", 0)],
        ["Average True Vertical Depth", summary.get("average_true_vertical_depth", 0)],
    ]

    for row in rows:
        summary_sheet.append(row)

    summary_sheet.column_dimensions["A"].width = 30
    summary_sheet.column_dimensions["B"].width = 20

    for cell in summary_sheet["A"] + summary_sheet["B"]:
        cell.alignment = Alignment(vertical="center", horizontal="left")

    inventory_sheet = workbook.create_sheet("Well Inventory")

    inventory_headers = [
        "Code",
        "Name",
        "Field",
        "Operator",
        "Well Type",
        "Status",
        "Spud Date",
        "Completion Date",
        "First Production Date",
        "Total Depth",
        "True Vertical Depth",
        "Tubing Size",
        "Casing Size",
        "Artificial Lift",
        "Reservoir",
        "Formation",
        "Latitude",
        "Longitude",
        "Active",
    ]

    inventory_sheet.append(inventory_headers)
    for cell in inventory_sheet[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    for item in report_data.get("history", []):
        inventory_sheet.append([
            item.get("code", ""),
            item.get("name", ""),
            item.get("field", ""),
            item.get("operator", ""),
            item.get("well_type", ""),
            item.get("status", ""),
            item.get("spud_date", ""),
            item.get("completion_date", ""),
            item.get("first_production_date", ""),
            item.get("total_depth", ""),
            item.get("true_vertical_depth", ""),
            item.get("tubing_size", ""),
            item.get("casing_size", ""),
            item.get("artificial_lift", ""),
            item.get("reservoir", ""),
            item.get("formation", ""),
            item.get("latitude", ""),
            item.get("longitude", ""),
            "Yes" if item.get("is_active") else "No",
        ])

    widths = [18, 28, 25, 25, 18, 18, 15, 15, 18, 18, 20, 18, 18, 20, 20, 20, 15, 15, 12]
    for index, width in enumerate(widths, start=1):
        column = inventory_sheet.cell(row=1, column=index).column_letter
        inventory_sheet.column_dimensions[column].width = width

    for row in inventory_sheet.iter_rows(min_row=2, max_col=len(inventory_headers)):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    def add_distribution_sheet(title, items, headers):
        sheet = workbook.create_sheet(title)
        sheet.append(headers)
        sheet["A1"].font = Font(bold=True)
        sheet["B1"].font = Font(bold=True)
        for item in items:
            sheet.append([item.get(headers[0].lower().replace(" ", "_"), ""), item.get(headers[1].lower().replace(" ", "_"), 0)])
        sheet.column_dimensions["A"].width = 30
        sheet.column_dimensions["B"].width = 18
        for row in sheet.iter_rows(min_row=2, max_col=2):
            for cell in row:
                cell.alignment = Alignment(horizontal="left")

    add_distribution_sheet(
        "By Type",
        report_data.get("well_type_distribution", []),
        ["Type", "Count"],
    )
    add_distribution_sheet(
        "By Status",
        report_data.get("status_distribution", []),
        ["Status", "Count"],
    )
    add_distribution_sheet(
        "By Field",
        report_data.get("field_distribution", []),
        ["Field", "Count"],
    )
    add_distribution_sheet(
        "By Operator",
        report_data.get("operator_distribution", []),
        ["Operator", "Count"],
    )
    add_distribution_sheet(
        "Artificial Lift",
        report_data.get("artificial_lift_distribution", []),
        ["Artificial Lift", "Count"],
    )

    return workbook
