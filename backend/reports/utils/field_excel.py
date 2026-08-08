from openpyxl import Workbook
from openpyxl.styles import Alignment, Font


def _style_sheet(sheet, widths):
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for index, width in enumerate(widths, start=1):
        column = sheet.cell(row=1, column=index).column_letter
        sheet.column_dimensions[column].width = width
    sheet.freeze_panes = "A2"


def _distribution_sheet(workbook, title, rows, key):
    sheet = workbook.create_sheet(title)
    sheet.append([key.replace("_", " ").title(), "Count"])
    for row in rows:
        sheet.append([row.get(key, ""), row.get("count", 0)])
    _style_sheet(sheet, [32, 16])


def create_field_workbook(report_data):
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Summary"
    summary_sheet.append(["Metric", "Value"])

    for key, label in [
        ("total_fields", "Total Fields"),
        ("active_fields", "Active Fields"),
        ("total_wells", "Total Wells"),
        ("producing_wells", "Producing Wells"),
        ("total_operators", "Total Operators"),
        ("total_oil", "Total Oil"),
        ("total_gas", "Total Gas"),
        ("total_water", "Total Water"),
    ]:
        summary_sheet.append([label, report_data.get("summary", {}).get(key, 0)])
    _style_sheet(summary_sheet, [30, 20])

    inventory = workbook.create_sheet("Field Inventory")
    inventory.append([
        "ID", "Field", "Code", "Operator", "Country", "State", "City",
        "Latitude", "Longitude", "Status", "Well Count", "Producing Wells",
        "Shut In Wells", "Drilling Wells",
    ])
    for item in report_data.get("history", []):
        inventory.append([
            item.get("id", ""), item.get("name", ""), item.get("code", ""),
            item.get("operator", ""), item.get("country", ""), item.get("state", ""),
            item.get("city", ""), item.get("latitude", ""), item.get("longitude", ""),
            item.get("status", ""), item.get("well_count", 0),
            item.get("producing_wells", 0), item.get("shut_in_wells", 0),
            item.get("drilling_wells", 0),
        ])
    _style_sheet(inventory, [8, 28, 16, 24, 18, 18, 18, 14, 14, 16, 14, 18, 16, 16])

    _distribution_sheet(workbook, "By Status", report_data.get("status_distribution", []), "status")
    _distribution_sheet(workbook, "By Operator", report_data.get("operator_distribution", []), "operator")
    if report_data.get("location_distribution"):
        _distribution_sheet(workbook, "By Location", report_data["location_distribution"], "location")
    _distribution_sheet(workbook, "Wells by Field", report_data.get("well_distribution", []), "field")
    _distribution_sheet(workbook, "Well Types", report_data.get("well_type_distribution", []), "well_type")

    if report_data.get("production_by_field"):
        production = workbook.create_sheet("Production by Field")
        production.append(["Field", "Oil", "Gas", "Water"])
        for item in report_data["production_by_field"]:
            production.append([
                item.get("field", ""), item.get("oil", 0),
                item.get("gas", 0), item.get("water", 0),
            ])
        _style_sheet(production, [32, 18, 18, 18])

    return workbook
