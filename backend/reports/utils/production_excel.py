from openpyxl import Workbook
from openpyxl.styles import Alignment, Font


def _style_sheet(sheet, widths):
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for index, width in enumerate(widths, start=1):
        column = sheet.cell(row=1, column=index).column_letter
        sheet.column_dimensions[column].width = width

    sheet.freeze_panes = "A2"


def create_production_workbook(report_data):
    workbook = Workbook()

    summary_sheet = workbook.active
    summary_sheet.title = "Summary"
    summary_sheet.append(["Metric", "Value", "Unit"])

    summary = report_data.get("summary", {})
    summary_rows = [
        ["Total Oil", summary.get("total_oil", 0), "BOPD"],
        ["Total Gas", summary.get("total_gas", 0), "MSCFD"],
        ["Total Water", summary.get("total_water", 0), "BWPD"],
        ["Average Oil", summary.get("average_oil", 0), "BOPD"],
        ["Average Gas", summary.get("average_gas", 0), "MSCFD"],
        ["Average Water", summary.get("average_water", 0), "BWPD"],
        ["Production Records", summary.get("total_records", 0), ""],
    ]
    for row in summary_rows:
        summary_sheet.append(row)
    _style_sheet(summary_sheet, [28, 18, 14])

    daily_sheet = workbook.create_sheet("Daily Production")
    daily_sheet.append([
        "Date",
        "Oil Production (BOPD)",
        "Gas Production (MSCFD)",
        "Water Production (BWPD)",
    ])
    for item in report_data.get("daily_production", []):
        daily_sheet.append([
            item.get("date", ""),
            item.get("oil", 0),
            item.get("gas", 0),
            item.get("water", 0),
        ])
    _style_sheet(daily_sheet, [16, 24, 26, 28])

    wells_sheet = workbook.create_sheet("Top Wells")
    wells_sheet.append([
        "Well Code",
        "Well Name",
        "Oil Production (BOPD)",
        "Gas Production (MSCFD)",
        "Water Production (BWPD)",
    ])
    for item in report_data.get("top_wells", []):
        wells_sheet.append([
            item.get("well_code", ""),
            item.get("well_name", ""),
            item.get("oil", 0),
            item.get("gas", 0),
            item.get("water", 0),
        ])
    _style_sheet(wells_sheet, [18, 30, 24, 26, 28])

    return workbook
